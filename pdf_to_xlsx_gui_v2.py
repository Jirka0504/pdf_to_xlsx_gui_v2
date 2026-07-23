import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

import pdfplumber
import pytesseract
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font


APP_TITLE = "PDF → XLSX GUI v2 (ASWO + OMNIA)"


HEADERS = [
    "Kód zboží dodavatele",
    "Název",
    "Množství",
    "Cena celkem",
    "Zkrácená poznámka",
    "Kód kombinované nomenklatury",
    "Země původu",
    "Hmotnost",
]


def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("￾", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def cz_number_to_float(value: str) -> float:
    value = value.replace(" ", "").replace("\xa0", "").replace(".", "").replace(",", ".")
    return float(value)


def en_number_to_float(value: str) -> float:
    value = value.replace(" ", "").replace("\xa0", "").replace(",", "")
    return float(value)


def float_to_dot_string(value: float) -> str:
    return f"{float(value):.2f}"


def extract_pdf_lines(pdf_path: str):
    lines = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""

            for raw_line in text.splitlines():
                line = normalize_text(raw_line)

                if line:
                    lines.append(line)

    return lines

def detect_supplier(lines: list[str]) -> str:
    text = "\n".join(lines).lower()

    if "aswo czech s.r.o." in text or "číslo produktu" in text or "dodávané množství" in text:
        return "ASWO"

    # NOVÝ FORMÁT OMNIA - musí být PŘED starou OMNIA detekcí
    if (
        "code cust. item description qty. u.m. price amount vat c." in text
        and "omnia components srl" in text
    ):
        return "OMNIA_NEW"

    if "omnia components" in text or "product code description quantity prezzo" in text:
        return "OMNIA"

    if "classic service parts gmbh" in text or "invoice no." in text and "priceeur" in text and "totaleur" in text:
        return "CLASSIC"

    if "vor spol. s r.o." in text or "cenabezdph" in text or "dodací list" in text:
        return "VOR"

    if ("phono-zubehör-vertrieb gmbh" in text or "analogis" in text and "art.nr." in text and "gesamtpreis" in text):
        return "ANALOGIS"

    if ("eta a. s." in text and "artikl" in text and ("popis artiklu" in text or "název" in text)):
        return "ETA"

    return "UNKNOWN"


# =========================
# ASWO PARSER
# =========================

def looks_like_split_code_only(line: str) -> bool:
    line = normalize_text(line)
    return bool(re.fullmatch(r"[A-Z0-9\-_/\.]+", line))


def parse_aswo(lines: list[str]):
    item_main_re = re.compile(
        r"""
        ^
        (?P<code>[A-Z0-9\-_/\.]+)\s+
        (?P<name>.+?)\s+
        (?P<ordered>\d+)\s+
        (?P<delivered>\d+)\s+
        (?P<unit_sale>\d{1,3}(?:\.\d{3})*,\d{2})\s+
        (?P<unit_net>\d{1,3}(?:\.\d{3})*,\d{2})\s+
        (?P<vat>\d{1,2})\s+
        (?P<total>\d{1,3}(?:\.\d{3})*,\d{2})
        $
        """,
        re.VERBOSE,
    )

    details_re = re.compile(
        r"""
        ^
        Celní\ kód\ zboží:\s*(?P<customs>\d+)\s+
        Hmotnost\ v\ gramech:\s*(?P<weight>\d+)\s+
        Původ:\s*(?P<origin>[A-Z]{2,3})
        $
        """,
        re.VERBOSE,
    )

    job_re = re.compile(r"^Číslo zakázky:\s*(?P<job>.+)$")
    logistics_re = re.compile(r"^Plus poměrné logistické náklady\s*(?P<log>\d{1,3}(?:\.\d{3})*,\d{2})\s*CZK$")

    skip_prefixes = (
        "Stránka:",
        "Objednávka:",
        "Zákaznické číslo:",
        "Datum vystavení",
        "dokladu / DUZP",
        "Číslo dokladu:",
        "DIČ:",
        "Údaje o vaší objednávce:",
        "Objednávka z:",
        "Jméno zákazníka:",
        "Číslo objednávky:",
        "Poznámka k dodávce",
        "Dodací adresa:",
        "ASWO Czech s.r.o.",
        "Na Prosecké vyhlídce",
        "info@aswo.cz",
        "Výkonný ředitel:",
        "Krajský soud:",
        "Místo plnění a soudní příslušnost:",
        "Commerzbank AG",
        "Číslo účtu:",
        "Kód banky:",
        "Číslo produktu",
        "Objednané množství",
        "Dodávané množství",
        "Celková cena",
        "Označení produktu",
        "Prodejní cena za jednotku",
        "Čistá cena za jednotku",
        "Sazba daně %",
        "Čistá hodnota celkem",
        "Náklady na dopravu",
        "Hodnota zboží",
        "Příplatek za velký díl",
    )

    items = []
    warnings = []

    current_item = None
    pending_code_fragment = None

    def push_current():
        nonlocal current_item
        if current_item:
            items.append(current_item)
            current_item = None

    for line in lines:
        if any(line.startswith(prefix) for prefix in skip_prefixes):
            continue

        if line in ("KTS-AME S.R.O.", "Karla Capka 60", "50002 Hradec Kralove", "CZECH REPUBLIC", "Jiří Krejčí"):
            continue

        if pending_code_fragment:
            test_line = f"{pending_code_fragment} {line}"
            m = item_main_re.match(test_line)
            if m:
                push_current()
                current_item = {
                    "code": m.group("code").strip(),
                    "name": m.group("name").strip(),
                    "qty": int(m.group("delivered")),
                    "total_base": cz_number_to_float(m.group("total")),
                    "logistics": 0.0,
                    "job": "",
                    "customs": "",
                    "origin": "",
                    "weight": "",
                }
                pending_code_fragment = None
                continue

        m = item_main_re.match(line)
        if m:
            push_current()
            current_item = {
                "code": m.group("code").strip(),
                "name": m.group("name").strip(),
                "qty": int(m.group("delivered")),
                "total_base": cz_number_to_float(m.group("total")),
                "logistics": 0.0,
                "job": "",
                "customs": "",
                "origin": "",
                "weight": "",
            }
            continue

        if looks_like_split_code_only(line):
            pending_code_fragment = line
            continue

        if current_item:
            m2 = details_re.match(line)
            if m2:
                current_item["customs"] = m2.group("customs").strip()
                current_item["origin"] = ""
                current_item["weight"] = int(m2.group("weight"))
                continue

            m3 = job_re.match(line)
            if m3:
                current_item["job"] = m3.group("job").strip()
                continue

            m4 = logistics_re.match(line)
            if m4:
                current_item["logistics"] = cz_number_to_float(m4.group("log"))
                continue

    push_current()

    result = []
    for item in items:
        total_with_logistics = item["total_base"] + item["logistics"]
        result.append({
            "Kód zboží dodavatele": item["code"],
            "Název": item["name"],
            "Množství": item["qty"],
            "Cena celkem": total_with_logistics,
            "Zkrácená poznámka": item["job"],
            "Kód kombinované nomenklatury": item["customs"],
            "Země původu": item["origin"],
            "Hmotnost": item["weight"],
        })

    return result, warnings


# =========================
# OMNIA PARSER
# =========================

def parse_omnia(lines: list[str]):
    items = []
    warnings = []

    full_row_re = re.compile(
        r"^(?P<code>[A-Z0-9\-.]+)\s+(?P<name>.+?)\s+(?P<qty>\d+)\s+PZ\s+(?P<price>\d+(?:\.\d{2}))\s+€\s+(?P<total>\d+(?:\.\d{2}))\s+€$"
    )

    name_qty_re = re.compile(
        r"^(?P<name>.+?)\s+(?P<qty>\d+)\s+PZ\s+(?P<price>\d+(?:\.\d{2}))\s+€\s+(?P<total>\d+(?:\.\d{2}))\s+€$"
    )

    skip_prefixes = (
        "Consegnare a:", "Spettabile:", "Vat code:", "TOTALE MERCE",
        "SCONTO %", "SPESE INCASSO", "TOTALE IMPONIBILE", "TOTALE IMPOSTA",
        "IMPONIBILE", "ALIQUOTA IVA", "TOTALE DOCUMENTO", "OMAGGI",
        "TOTALE DA PAGARE", "FATTURA", "OMNIA COMPONENTS", "Via Travnik",
        "Tel.", "C.F. E P.IVA", "Capitale Sociale", "N.Documento",
        "Data spedizione richiesta:", "http://www.omniacomponents.com",
        "PRODUCT CODE DESCRIPTION QUANTITY PREZZO SCONTO % IMPORTO TOTALE",
        "Spese di Trasporto UE Vendita - Shipping Fees",
    )

    cleaned_lines = []
    for raw in lines:
        line = str(raw).replace("￾", "-")
        line = normalize_text(line)
        if not line:
            continue
        if any(line.startswith(prefix) for prefix in skip_prefixes):
            continue
        if line.startswith("KTS - AME") or line.startswith("Karla Čapka") or line.startswith("500 02"):
            continue
        cleaned_lines.append(line)

    i = 0
    while i < len(cleaned_lines):
        line = cleaned_lines[i]

        # 1) běžný kompletní řádek
        m = full_row_re.match(line)
        if m:
            items.append({
                "Kód zboží dodavatele": m.group("code").strip(),
                "Název": m.group("name").strip(),
                "Množství": int(m.group("qty")),
                "Cena celkem": en_number_to_float(m.group("total")),
                "Zkrácená poznámka": "",
                "Kód kombinované nomenklatury": "",
                "Země původu": "",
                "Hmotnost": "",
            })
            i += 1
            continue

        # 2) třířádkový rozpad:
        # SS-
        # .3D/CROCHET/NOIR 1 PZ 7.02 € 7.02 €
        # 1600007234
        if re.fullmatch(r"[A-Z]{2,10}-$", line):
            if i + 2 < len(cleaned_lines):
                middle_line = cleaned_lines[i + 1]
                tail_code_line = cleaned_lines[i + 2]

                m2 = name_qty_re.match(middle_line)

                if m2 and re.fullmatch(r"[A-Z0-9.\-]+", tail_code_line):
                    full_code = line + tail_code_line

                    items.append({
                        "Kód zboží dodavatele": full_code,
                        "Název": m2.group("name").strip(),
                        "Množství": int(m2.group("qty")),
                        "Cena celkem": en_number_to_float(m2.group("total")),
                        "Zkrácená poznámka": "",
                        "Kód kombinované nomenklatury": "",
                        "Země původu": "",
                        "Hmotnost": "",
                    })
                    i += 3
                    continue

        # 3) dvouřádkový rozpad:
        # SS-1600007234
        # .3D/CROCHET/NOIR 1 PZ 7.02 € 7.02 €
        if re.fullmatch(r"[A-Z]{2,10}-[A-Z0-9.\-]+", line):
            if i + 1 < len(cleaned_lines):
                m3 = name_qty_re.match(cleaned_lines[i + 1])
                if m3:
                    items.append({
                        "Kód zboží dodavatele": line.strip(),
                        "Název": m3.group("name").strip(),
                        "Množství": int(m3.group("qty")),
                        "Cena celkem": en_number_to_float(m3.group("total")),
                        "Zkrácená poznámka": "",
                        "Kód kombinované nomenklatury": "",
                        "Země původu": "",
                        "Hmotnost": "",
                    })
                    i += 2
                    continue

        i += 1

    return items, warnings


# =========================
# OMNIA NEW PARSER
# =========================
   
def parse_omnia_new(lines: list[str]):
    items = []
    warnings = []

    full_row_re = re.compile(
        r"""
        ^
        (?P<code>[A-Z0-9.\-]+)\s+
        (?P<name>.+?)\s+
        (?P<qty>\d+)\s+
        Pcs\s+
        (?P<price>\d+(?:[.,]\d{2}))\s+
        (?P<total>\d+(?:[.,]\d{2}))\s+
        (?P<vat>\d+)
        $
        """,
        re.VERBOSE | re.IGNORECASE,
    )

    tail_re = re.compile(
        r"""
        ^
        (?P<qty>\d+)\s+
        Pcs\s+
        (?P<price>\d+(?:[.,]\d{2}))\s+
        (?P<total>\d+(?:[.,]\d{2}))\s+
        (?P<vat>\d+)
        $
        """,
        re.VERBOSE | re.IGNORECASE,
    )

    skip_prefixes = (
        "Code Cust. Item",
        "Shipment No.",
        "Note:",
        "Company Stamp",
        "Parcel units:",
        "Gross weight",
        "Goods aspect:",
        "Vat C.",
        "Incoterm:",
        "Shipment due to:",
        "Shipping Agent:",
        "Subscr. No.:",
        "Truck. No.:",
        "Shipping Starting Date",
        "Driver signature:",
        "Addressee signature:",
        "SWIFT - Bank transfer",
        "No. ",
        "Via Travnik",
        "Company subject",
        "SDI Code:",
        "VAT Registration",
        "Ph:",
        "Omnia Components Srl",
        "KTS - AME",
        "Czech Republic",
        "Company:",
        "Karla Čapka",
        "Deliver to:",
        "Invoice",
        "Payment:",
        "Paym. Method:",
        "Net 60",
        "Bill-to Customer",
        "Due Dates:",
        "Bank:",
        "PAG",
        "TOTAL DOCUMENT",
        "Pursuant to",
        "Delivery At Place",
        "Fedex",
    )

    cleaned = []

    for raw in lines:
        line = normalize_text(str(raw))

        if not line:
            continue

        if any(line.startswith(prefix) for prefix in skip_prefixes):
            continue

        cleaned.append(line)

    i = 0

    while i < len(cleaned):
        line = cleaned[i]

        # Dopravu vůbec neexportovat
        if line.startswith("TRASP.EU.VEN") or "Shipping Fees" in line:
            i += 1
            continue

        # 1) Běžná položka na jednom řádku
        m = full_row_re.match(line)

        if m:
            code = m.group("code").strip()

            if code == "TRASP.EU.VEN":
                i += 1
                continue

            items.append({
                "Kód zboží dodavatele": code,
                "Název": m.group("name").strip(),
                "Množství": int(m.group("qty")),
                "Cena celkem": float(
                    m.group("total").replace(",", ".")
                ),
                "Zkrácená poznámka": "",
                "Kód kombinované nomenklatury": "",
                "Země původu": "",
                "Hmotnost": "",
            })

            i += 1
            continue

        # 2) Položka s názvem rozděleným do více řádků
        start = re.match(
            r"^(?P<code>[A-Z0-9.\-]+)\s+(?P<name>.+)$",
            line,
        )

        if start:
            code = start.group("code").strip()

            if code == "TRASP.EU.VEN":
                i += 1
                continue

            name_parts = [start.group("name").strip()]
            j = i + 1
            found = False

            while j < len(cleaned):
                nxt = cleaned[j]

                # Hledáme řádek obsahující:
                # množství + Pcs + jednotkovou cenu + celkovou cenu + VAT
                tail = tail_re.match(nxt)

                if tail:
                    items.append({
                        "Kód zboží dodavatele": code,
                        "Název": " ".join(name_parts).strip(),
                        "Množství": int(tail.group("qty")),
                        "Cena celkem": float(
                            tail.group("total").replace(",", ".")
                        ),
                        "Zkrácená poznámka": "",
                        "Kód kombinované nomenklatury": "",
                        "Země původu": "",
                        "Hmotnost": "",
                    })

                    i = j + 1
                    found = True
                    break

                # Pokud narazíme na dopravu, položku ukončíme
                if nxt.startswith("TRASP.EU.VEN") or "Shipping Fees" in nxt:
                    break

                # Pokud narazíme na další kompletní položku,
                # nepřipojujeme ji k názvu předchozí položky
                if full_row_re.match(nxt):
                    break

                # Konec tabulky / souhrny faktury
                if nxt.startswith((
                    "Übertrag",
                    "Nettobetrag",
                    "Mehrwertsteuer",
                    "Bruttobetrag",
                    "Zahlungsbedingungen",
                    "Soweit nicht anders",
                    "TOTAL DOCUMENT",
                )):
                    break

                name_parts.append(nxt.strip())
                j += 1

            if found:
                continue

        i += 1

    return items, warnings
    

# =========================
# CLASSIC PARSER
# =========================

def parse_classic(lines: list[str]):
    items = []
    warnings = []

    cleaned_lines = []

    skip_words = [
        "invoice",
        "classic service parts",
        "karla capka",
        "tschechische",
        "tracking",
        "dpd",
        "vat",
        "goods value",
        "total amount",
        "payment within",
        "grossweight",
        "eur",
    ]

    for raw in lines:
        line = normalize_text(raw)
        if not line:
            continue

        low = line.lower()

        if any(word in low for word in skip_words):
            continue

        # oddělovače
        if "====" in line:
            continue

        cleaned_lines.append(line)

    # hlavní řádek položky
    item_re = re.compile(
        r"^(\d{6,})\s+(\d+,\d{2})\s+(.+?)\s+(\d+,\d{2})\s+(\d+,\d{2})$"
    )

    i = 0
    while i < len(cleaned_lines):
        line = cleaned_lines[i]
        m = item_re.match(line)

        if not m:
            i += 1
            continue

        qty = cz_number_to_float(m.group(2))
        name = m.group(3).strip()
        total = cz_number_to_float(m.group(5))

        internal_code = ""

        # 🔑 druhý řádek = správný kód (BLTxxxxx)
        if i + 1 < len(cleaned_lines):
            next_line = cleaned_lines[i + 1].strip()

            # musí obsahovat písmena + čísla
            if re.match(r"^[A-Z0-9\- ]+$", next_line):
                internal_code = next_line
                i += 1

        # ❌ vynechání dopravy
        if "FRA" in internal_code or "FREIGHT" in name.upper():
            i += 1
            continue

        items.append({
            "Kód zboží dodavatele": internal_code,
            "Název": name,
            "Množství": qty,
            "Cena celkem": total,
            "Zkrácená poznámka": "",
            "Kód kombinované nomenklatury": "",
            "Země původu": "",
            "Hmotnost": "",
        })

        i += 1

    return items, warnings  

# =========================
# VOR PARSER
# =========================

def parse_vor(lines: list[str]):
    items = []
    warnings = []

    item_start_re = re.compile(r"^\d{2}\s+\S+\s+\d+\s+ks\s+")

    item_re = re.compile(
        r"""
        ^
        (?P<code>\d{2}\s+\S+)\s+
        (?P<qty>\d+)\s+ks\s+
        (?P<body>.+?)
        \s+
        (?P<unit_price>\d+(?:\.\d{3}))
        \s+
        (?P<vat>\d{1,2})
        \s+
        (?P<vat_amount>\d+(?:,\d{3})*\.\d{2})
        \s+
        (?P<total_vat>\d+(?:,\d{3})*\.\d{2})\s*Kč?
        (?P<after>.*)
        $
        """,
        re.VERBOSE,
    )

    stop_re = re.compile(
        r"^(Po zaokrouhlení|Základní sazba|Zaplaceno|Celkem k úhradě|V cenové skupině|Za obaly|Kontrola expedice|Datum:|Fakturu vystavil|doklad č\.)"
    )

    skip_prefixes = (
        "Daňový doklad",
        "Variabilní symbol",
        "DODACÍ LIST",
        "Pokud nebyla",
        "Prodávající",
        "Kupující",
        "Banka:",
        "Účet/Kód banky:",
        "SWIFT:",
        "IBAN:",
        "Komerční banka",
        "KTS - AME",
        "Karla Čapka",
        "Kosice",
        "Dodací adresa:",
        "Datum UZP:",
        "Datum vystavení:",
        "Datum splatnosti:",
        "Způsob platby:",
        "Doprava:",
        "Zboží",
        "/MJ",
        "KUPNÍ SMLOUVA",
        "strana :",
        "1026565377",
        "20,444.66 Kč",
    )

    cleaned = []
    for raw in lines:
        line = normalize_text(str(raw))
        if not line:
            continue
        if any(line.startswith(p) for p in skip_prefixes):
            continue
        if stop_re.match(line):
            continue
        cleaned.append(line)

    i = 0
    while i < len(cleaned):
        line = cleaned[i]

        if not item_start_re.match(line):
            i += 1
            continue

        combined = line
        j = i + 1

        while j < len(cleaned):
            nxt = cleaned[j]

            if item_start_re.match(nxt):
                break
            if stop_re.match(nxt):
                break

            combined += " " + nxt
            j += 1

        m = item_re.match(combined)

        if not m:
            warnings.append(f"VOR: nepodařilo se naparsovat položku: {combined}")
            i = j
            continue

        code = m.group("code").strip()
        qty = int(m.group("qty"))
        unit_price = float(m.group("unit_price"))
        total = qty * unit_price

        body = m.group("body").strip()
        after = m.group("after").strip()

        full_text = (body + " " + after).strip()

        # název = vše před prvním //
        name = full_text.split("//", 1)[0].strip()

        items.append({
            "Interní kód zboží": code,
            "Kód zboží dodavatele": code,
            "Název": name,
            "Množství": qty,
            "Cena celkem": total,
            "Zkrácená poznámka": "",
            "Kód kombinované nomenklatury": "",
            "Země původu": "",
            "Hmotnost": "",
        })

        i = j

    return items, warnings

# =========================
# ANALOGIS PARSER
# =========================

def parse_analogis(lines: list[str]):
    items = []
    warnings = []

    row_re = re.compile(
        r"""
        ^
        (?P<pos>\d+)\s+
        (?P<code>\S+)\s+
        (?P<name>.+?)\s+
        (?P<qty>\d+)\s+
        (?P<price>\d+(?:,\d{2}))\s*€\s+
        (?P<total>\d+(?:,\d{2}))\s*€
        $
        """,
        re.VERBOSE,
    )

    skip_prefixes = (
        "Pos. Art.Nr.",
        "Phono-Zubehör-Vertrieb",
        "Fon:",
        "Eingetragen:",
        "Amtsgericht",
        "Geschäftsführer:",
        "UST-",
        "Steuernummer:",
        "WEEE-",
        "Duales System:",
        "Es gelten",
        "Jana Kallweit",
        "KTS - AME",
        "Karla Capka",
        "50002",
        "TSCHECHISCHE",
        "BEI ZAHLUNG",
        "Rechnung",
        "Rechnungsnummer:",
        "Sachbearbeiter:",
        "Datum:",
        "Lieferscheinnr.",
        "Ihre Kundennummer:",
        "Umsatzsteuerfreie",
        "Seite:",
        "Übertrag",
        "Soweit nicht anders",
        "Zahlungsbedingungen",
        "Nettobetrag",
        "Mehrwertsteuer",
        "Bruttobetrag",
        "Die Lieferung ist",
        "Ihre UstID",
    )

    cleaned = []

    for raw in lines:
        line = normalize_text(str(raw))
        if not line:
            continue

        if any(line.startswith(prefix) for prefix in skip_prefixes):
            continue

        cleaned.append(line)

    i = 0

    while i < len(cleaned):
        line = cleaned[i]
        m = row_re.match(line)

        if not m:
            i += 1
            continue

        code = m.group("code").strip()
        qty = int(m.group("qty"))
        total = float(m.group("total").replace(",", "."))
        name_parts = [m.group("name").strip()]

        # dopravu neexportovat
        if code.upper() == "DHL-PAKET" or "DHL-PARCEL" in line.upper():
            i += 1
            continue

        # nedodané položky vynechat
        if qty == 0:
            i += 1
            continue

        j = i + 1

        while j < len(cleaned):
            nxt = cleaned[j]

            # další položka
            if row_re.match(nxt):
                break

            # konec tabulky / souhrny
            if nxt.startswith((
                "Übertrag",
                "Nettobetrag",
                "Mehrwertsteuer",
                "Bruttobetrag",
                "Zahlungsbedingungen",
                "Soweit nicht anders",
            )):
                break

            name_parts.append(nxt.strip())
            j += 1

        items.append({
            "Kód zboží dodavatele": code,
            "Název": " ".join(name_parts).strip(),
            "Množství": qty,
            "Cena celkem": total,
            "Zkrácená poznámka": "",
            "Kód kombinované nomenklatury": "",
            "Země původu": "",
            "Hmotnost": "",
        })

        i = j

    return items, warnings

# =========================
# ETA PARSER
# =========================

def parse_eta(lines: list[str]):
    items = []
    warnings = []

    price_pattern = r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}"

    row_re = re.compile(
        rf"""
        ^
        (?P<code>ETA\d+|\#\d+)\s+
        (?P<name>.+?)\s+
        (?P<qty>\d+)\s+
        (?P<vat>\d+)\s+
        (?P<unit_price>{price_pattern})\s+
        (?P<unit_price_rp>{price_pattern})\s+
        (?P<total>{price_pattern})\s+
        (?P<vat_total>{price_pattern})
        $
        """,
        re.VERBOSE | re.IGNORECASE,
    )

    tail_re = re.compile(
        rf"""
        ^
        (?P<qty>\d+)\s+
        (?P<vat>\d+)\s+
        (?P<unit_price>{price_pattern})\s+
        (?P<unit_price_rp>{price_pattern})\s+
        (?P<total>{price_pattern})\s+
        (?P<vat_total>{price_pattern})
        $
        """,
        re.VERBOSE,
    )

    item_start_re = re.compile(
        r"^(?P<code>ETA\d+|\#\d+)\s+(?P<name>.+)$",
        re.IGNORECASE,
    )

    skip_prefixes = (
        "Číslo zakázky:",
        "Externí číslo obj.:",
        "Dodavatel:",
        "Zákazník:",
        "ETA a. s.",
        "Křižíkova",
        "18600 Praha",
        "IČO:",
        "Tel.:",
        "Email:",
        "KTS - AME",
        "Karla Čapka",
        "50002 Hradec",
        "Datum vystavení:",
        "Datum odeslání:",
        "Způsob dopravy:",
        "Forma úhrady:",
        "Dodací adresa:",
        "Kosice",
        "503 51",
        "Artikl Popis artiklu",
        "Artikl",
        "Sazba % DPH",
        "Celkem bez RP:",
        "Celkem s RP:",
        "Celkem k úhradě:",
        "Celkové množství",
        "Celková hmotnost",
        "Celkový objem",
        "Kalkulace ceny",
    )

    def eta_price_to_float(value: str) -> float:
        value = str(value).strip()
        value = value.replace("\xa0", "")
        value = value.replace(" ", "")
        value = value.replace(".", "")
        value = value.replace(",", ".")
        return float(value)

    def eta_clean_code(code: str) -> str:
        code = str(code).strip()

        if code.upper().startswith("ETA"):
            code = code[3:]

        if code.startswith("#"):
            code = code[1:]

        return code.strip()

    cleaned = []

    for raw in lines:
        line = normalize_text(str(raw))

        if not line:
            continue

        if any(line.startswith(prefix) for prefix in skip_prefixes):
            continue

        if re.fullmatch(r"\d+\s*/\s*\d+", line):
            continue

        cleaned.append(line)

    i = 0

    while i < len(cleaned):
        line = cleaned[i]

        m = row_re.match(line)

        if m:
            items.append({
                "Kód zboží dodavatele": eta_clean_code(m.group("code")),
                "Název": m.group("name").strip(),
                "Množství": int(m.group("qty")),
                "Cena celkem": eta_price_to_float(m.group("total")),
                "Zkrácená poznámka": "",
                "Kód kombinované nomenklatury": "",
                "Země původu": "",
                "Hmotnost": "",
            })

            i += 1
            continue

        start = item_start_re.match(line)

        if start:
            original_code = start.group("code").strip()
            name_parts = [start.group("name").strip()]

            j = i + 1
            found = False

            while j < len(cleaned):
                nxt = cleaned[j]

                tail = tail_re.match(nxt)

                if tail:
                    items.append({
                        "Kód zboží dodavatele": eta_clean_code(original_code),
                        "Název": " ".join(name_parts).strip(),
                        "Množství": int(tail.group("qty")),
                        "Cena celkem": eta_price_to_float(tail.group("total")),
                        "Zkrácená poznámka": "",
                        "Kód kombinované nomenklatury": "",
                        "Země původu": "",
                        "Hmotnost": "",
                    })

                    i = j + 1
                    found = True
                    break

                combined_name = " ".join(name_parts + [nxt]).strip()
                fake_line = f"{original_code} {combined_name}"

                m2 = row_re.match(fake_line)

                if m2:
                    items.append({
                        "Kód zboží dodavatele": eta_clean_code(m2.group("code")),
                        "Název": m2.group("name").strip(),
                        "Množství": int(m2.group("qty")),
                        "Cena celkem": eta_price_to_float(m2.group("total")),
                        "Zkrácená poznámka": "",
                        "Kód kombinované nomenklatury": "",
                        "Země původu": "",
                        "Hmotnost": "",
                    })

                    i = j + 1
                    found = True
                    break

                if item_start_re.match(nxt):
                    break

                if nxt.startswith((
                    "Sazba % DPH",
                    "Celkem bez RP:",
                    "Celkem s RP:",
                    "Celkem k úhradě:",
                    "Celkové množství",
                    "Celková hmotnost",
                    "Celkový objem",
                    "Kalkulace ceny",
                )):
                    break

                name_parts.append(nxt.strip())
                j += 1

            if found:
                continue

            warnings.append(
                f"ETA: nepodařilo se naparsovat položku: {line}"
            )

        i += 1

    return items, warnings
    
# =========================
# SAVE XLSX
# =========================

def save_xlsx(output_path: str, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "List1"

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    row_no = 2
    for row in rows:
        ws.cell(row=row_no, column=1, value=row["Kód zboží dodavatele"])
        ws.cell(row=row_no, column=2, value=row["Název"])
        ws.cell(row=row_no, column=3, value=row["Množství"])
        ws.cell(row=row_no, column=4, value=float_to_dot_string(row["Cena celkem"]))
        ws.cell(row=row_no, column=5, value=row["Zkrácená poznámka"])
        ws.cell(row=row_no, column=6, value=int(row["Kód kombinované nomenklatury"]) if str(row["Kód kombinované nomenklatury"]).isdigit() else row["Kód kombinované nomenklatury"])
        ws.cell(row=row_no, column=7, value=row["Země původu"])
        ws.cell(row=row_no, column=8, value=row["Hmotnost"])
        row_no += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12

    wb.save(output_path)


# =========================
# GUI
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("980x640")

        self.pdf_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.detected_supplier = tk.StringVar(value="Nezjištěno")

        self.preview_box = None
        self._build_ui()

    def _build_ui(self):
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Vstupní PDF:").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(frm, textvariable=self.pdf_path, width=95).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Button(frm, text="Vybrat…", command=self.pick_pdf).grid(row=0, column=2, padx=6)

        ttk.Label(frm, text="Výstupní XLSX:").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(frm, textvariable=self.output_path, width=95).grid(row=1, column=1, sticky="ew", pady=4)
        ttk.Button(frm, text="Uložit jako…", command=self.pick_output).grid(row=1, column=2, padx=6)

        ttk.Label(frm, text="Detekovaný typ PDF:").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(frm, textvariable=self.detected_supplier, width=30, state="readonly").grid(row=2, column=1, sticky="w", pady=4)

        ttk.Separator(frm, orient="horizontal").grid(row=3, column=0, columnspan=3, sticky="ew", pady=10)

        note = (
            "Verze 2:\n"
            "• Automatická detekce ASWO / OMNIA / CLASSIC\n"
            "• ASWO: množství = dodávané množství\n"
            "• ASWO: cena celkem = celková cena + poměrné logistické náklady\n"
            "• OMNIA: umí i rozdělený kód na dalším řádku\n"
            "• CLASSIC: umí 2–3 řádkové položky (Id-No. / druhý kód / poznámka)"
        )
        ttk.Label(frm, text=note, foreground="#555", justify="left").grid(row=4, column=0, columnspan=3, sticky="w", pady=(0, 10))

        self.preview_box = tk.Text(frm, height=22, wrap="none")
        self.preview_box.grid(row=5, column=0, columnspan=3, sticky="nsew")

        btns = ttk.Frame(frm)
        btns.grid(row=6, column=0, columnspan=3, sticky="e", pady=12)
        ttk.Button(btns, text="Detekovat", command=self.detect_only).pack(side="left", padx=4)
        ttk.Button(btns, text="Náhled", command=self.preview).pack(side="left", padx=4)
        ttk.Button(btns, text="Převést", command=self.convert).pack(side="left", padx=4)
        ttk.Button(btns, text="Konec", command=self.destroy).pack(side="left", padx=4)

        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(5, weight=1)

    def pick_pdf(self):
        path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if path:
            self.pdf_path.set(path)
            if not self.output_path.get().strip():
                self.output_path.set(str(Path(path).with_suffix(".xlsx")))

    def pick_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.output_path.set(path)

    def _load_and_detect(self):
        pdf = self.pdf_path.get().strip()
        if not pdf:
            raise ValueError("Vyber vstupní PDF.")

        lines = extract_pdf_lines(pdf)
        supplier = detect_supplier(lines)
        self.detected_supplier.set(supplier)
        return lines, supplier

    def detect_only(self):
        try:
            _, supplier = self._load_and_detect()
        except Exception as e:
            messagebox.showerror("Chyba", str(e))
            return
        messagebox.showinfo("Detekce", f"Rozpoznaný typ PDF: {supplier}")

    def _parse(self):
        lines, supplier = self._load_and_detect()

        if supplier == "ASWO":
            return parse_aswo(lines), supplier

        if supplier == "OMNIA_NEW":
            return parse_omnia_new(lines), supplier

        if supplier == "OMNIA":
            return parse_omnia(lines), supplier

        if supplier == "CLASSIC":
            return parse_classic(lines), supplier

        if supplier == "VOR":
            return parse_vor(lines), supplier

        if supplier == "ANALOGIS":
            return parse_analogis(lines), supplier

        if supplier == "ETA":
            return parse_eta(lines), supplier

        raise ValueError("Nepodařilo se rozpoznat typ PDF. Tento dokument zatím není podporovaný.")
        
    def preview(self):
        try:
            (rows, warnings), supplier = self._parse()
        except Exception as e:
            messagebox.showerror("Chyba", str(e))
            return

        self.preview_box.delete("1.0", "end")
        self.preview_box.insert("1.0", f"Typ PDF: {supplier}\n")
        self.preview_box.insert("end", f"Nalezeno položek: {len(rows)}\n\n")

        for row in rows[:40]:
            self.preview_box.insert(
                "end",
                f"{row['Kód zboží dodavatele']} | {row['Název']} | qty={row['Množství']} | total={float_to_dot_string(row['Cena celkem'])}\n"
            )

        if warnings:
            self.preview_box.insert("end", "\n--- VAROVÁNÍ ---\n")
            for w in warnings[:20]:
                self.preview_box.insert("end", w + "\n")

    def convert(self):
        out = self.output_path.get().strip()
        if not out:
            messagebox.showerror("Chyba", "Vyber výstupní XLSX.")
            return

        try:
            (rows, warnings), supplier = self._parse()
            save_xlsx(out, rows)
        except Exception as e:
            messagebox.showerror("Chyba", str(e))
            return

        msg = f"Hotovo.\nTyp PDF: {supplier}\nUloženo do:\n{out}\n\nPočet položek: {len(rows)}"
        if warnings:
            msg += f"\nVarování: {len(warnings)}"
        messagebox.showinfo("Hotovo", msg)


if __name__ == "__main__":
    App().mainloop()
